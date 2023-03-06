# MIT License
#
# Copyright (c) 2022 herrsmitty8128
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
# 
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
# 
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

import os
import sys
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import load_workbook


required_column_headers = [
    'Number',
    'Facility',
    'Aging',
    'FC Begin',
    'FC End',
    'Start Date',
    'End Date',
    'Begin Bal',
    'Charges',
    'Admin',
    'Bad Debt',
    'Charity',
    'Contractuals',
    'Denials',
    'Payments',
    'End Bal',
    'Rsv: Begin Bal',
    'Rsv: End Bal'
]


def get_csv_header_crosswalk(filename: str) -> dict:
    headers = dict()
    with open(filename, 'r') as file:
        for line in file.readlines():
            if line.strip().startswith('#'): continue
            vals = line.split('=')
            count = len(vals)
            if count == 0: continue
            if count != 2:
                print('################ FATAL ERROR!!! ################')
                print('An error was encountered on the following line while reading the column header crosswalk file.')
                print(line)
                _ = input('Press ENTER to continue...')
                exit()
            headers[vals[0].strip().replace('"','')] = vals[1].strip().replace('"','')
    required_headers = set(x for x in required_column_headers)
    if headers.keys() != required_headers:
        print('################ FATAL ERROR!!! ################')
        print('Not all required headers are included in the column header crosswalk file.')
        print('The following headers are missing from the first column:')
        for d in required_headers.difference(headers.keys()):
            print('  ', d)
        _ = input('Press ENTER to continue...')
        exit()
    return headers


def serialize_date(csv_row: dict, field_name: str, header_lookup: dict) -> datetime:
    string = csv_row[header_lookup[field_name]].strip()
    try:
        return datetime.strptime(string.strip(), '%m/%d/%Y')
    except:
        raise ValueError('Invalid date string encountered.')


def serialize_currency(csv_row: dict, field_name: str, header_lookup: dict) -> float:
    string = csv_row[header_lookup[field_name]].strip()
    if string == '-' or string == '':
        return 0.0
    string.replace('(', '-').replace(')', '').replace(',', '').replace('$', '')
    if '.' not in string:
        string += '.0'
    return round(float(string), 2)


def load_input_files(directory: str, filenames: list[str], header_lookup: dict) -> list[dict]:
    data = []
    for filename in filenames:
        input_file = directory + filename
        with open(input_file, 'r', newline='') as f:
            reader = csv.DictReader(f)
            required_headers = set(h for h in header_lookup.values())
            if not required_headers.issubset(set(reader.fieldnames)):
                print('################ FATAL ERROR!!! ################')
                print('File', filename, 'is missing the following headers:')
                for d in required_headers.difference(set(reader.fieldnames)):
                    print('  ', d)
                _ = input('Press ENTER to continue...')
                exit()
            for row in reader:
                data.append({
                    'Number': int(row[header_lookup['Number']].strip()),
                    'Facility': row[header_lookup['Facility']].strip(),
                    'Aging': row[header_lookup['Aging']].strip(),
                    'FC Begin': row[header_lookup['FC Begin']].strip(),
                    'FC End': row[header_lookup['FC End']].strip(),
                    'Start Date': serialize_date(row, 'Start Date', header_lookup),
                    'End Date': serialize_date(row, 'End Date', header_lookup),
                    'Begin Bal': serialize_currency(row, 'Begin Bal', header_lookup),
                    'Charges': serialize_currency(row, 'Charges', header_lookup),
                    'Admin': serialize_currency(row, 'Admin', header_lookup),
                    'Bad Debt': serialize_currency(row, 'Bad Debt', header_lookup),
                    'Charity': serialize_currency(row, 'Charity', header_lookup),
                    'Contractuals': serialize_currency(row, 'Contractuals', header_lookup),
                    'Denials': serialize_currency(row, 'Denials', header_lookup),
                    'Payments': serialize_currency(row, 'Payments', header_lookup),
                    'End Bal': serialize_currency(row, 'End Bal', header_lookup),
                    'Rsv: Begin Bal': -serialize_currency(row, 'Rsv: Begin Bal', header_lookup),
                    'Rsv: End Bal': -serialize_currency(row, 'Rsv: End Bal', header_lookup)
                })
    return data


def add_reserve_activity(patients: list[dict]) -> None:
    # a list of the field names for which we'd like to calcualte reserve activity
    fields = ['Charges', 'Admin', 'Bad Debt', 'Charity', 'Contractuals', 'Denials']
    for pt in patients:
        # initialize the cumulative balances
        cum_acct_bal = pt['Begin Bal']
        cum_rsv_bal = pt['Rsv: Begin Bal']
        # calculate the reserve activity and add it to the patient account
        for field in fields:
            rsv_field = 'Rsv: ' + field
            pt[rsv_field] = 0.0 if cum_acct_bal == 0.0 else round(pt[field] / cum_acct_bal * cum_rsv_bal, 2)
            cum_acct_bal += pt[field]
            cum_rsv_bal += pt[rsv_field]
        # calculate the change in the reserve due the the month-end valuation
        pt['Rsv: Releases'] = 0.0
        pt['Rsv: Valuation'] = round(pt['Rsv: End Bal'] - cum_rsv_bal, 2)
        # verify that the reserve rolls-forward correctly
        if round(cum_rsv_bal + pt['Rsv: Valuation'], 2) != round(pt['Rsv: End Bal'], 2):
            raise ValueError(f'Reserve activity does not roll-forward for patient {pt["Number"]}')
        # relcassify the valuation amount if it is attributable to cash receipts
        if pt['Rsv: Begin Bal'] < 0.0 and pt['Rsv: End Bal'] == 0.0 and pt['Payments'] < 0.0 and pt['Rsv: Valuation'] > 0.0:
            pt['Rsv: Releases'] = pt['Rsv: Valuation']
            pt['Rsv: Valuation'] = 0.0
        pt['NPSR Impact'] = ((pt['End Bal'] - pt['Payments']) - pt['Begin Bal']) + (pt['Rsv: End Bal'] - pt['Rsv: Begin Bal'])
    pass


def add_descriptions(patients: list[dict]) -> tuple:
    for pt in patients:

        bd = pt['Bad Debt']
        beg_rsv = pt['Rsv: Begin Bal']
        beg_net = pt['Begin Bal'] + beg_rsv
        pay = pt['Payments']
        chg = pt['Charges']
        end_net = pt['End Bal'] + pt['Rsv: End Bal']

        desc = ''

        if beg_net > 0.0:
            desc += 'Positive beginning NRV '
        elif beg_net < 0.0:
            desc += 'Negative beginning NRV '
        else:
            desc += 'Zero beginning NRV '

        if chg > 0.0:
            desc += 'having charges, '
        elif chg < 0.0:
            desc += 'having charge reversals, '
        else:
            desc += 'having charges, '

        if bd > 0.0:
            desc += 'bad debt reversal, '

        if pay < 0.0:
            desc += 'cash receipts'
            if abs(pay) > (beg_net + chg + pt['Rsv: Charges']):
                desc += ' in excess of beginning NRV + net charges'
            desc += ', '
        elif pay > 0.0:
            desc += 'refunds, '
        else:
            desc += 'no cash activity, '

        if end_net > 0.0:
            desc += 'and ending in a debit NRV.'
        elif end_net < 0.0:
            desc += 'and ending in a credit NRV.'
        else:
            desc += 'and ending in a zero NRV.'

        pt['Description'] = desc

    pass


def write_to_file(filename: str, patients: list[dict]) -> None:
    headers = [h for h in patients[0].keys()]
    descriptors = [{
        'table_headers': headers, #required_column_headers,
        'table_rows': patients,
        'table_name': 'PT_ACCT_ROLL',
        'sheet_name': 'Pt Acct Roll-forward'
    }]
    wb = Workbook()
    active_sheet_name = wb.active.title
    for d in descriptors:
        new_table(wb.create_sheet(d['sheet_name']), d['table_headers'], d['table_rows'], d['table_name'])
    del wb[active_sheet_name]
    wb.save(filename)


def delete_all_rows(ws: Worksheet) -> None:
    '''
    Deletes all the rows in a worksheet.
    '''
    ws.delete_rows(1, ws.max_row)


def new_table(ws: Worksheet, table_headers: list[str], table_rows: list[dict], table_name: str) -> None:
    # Be sure to first delete the existing rows if you want the table to start on row 1
    delete_all_rows(ws)
    ws.append(table_headers)
    # add rows of data
    for row in table_rows:
        if isinstance(row, dict):
            ws.append([row[h] for h in table_headers])
        else:
            ws.append(row)
    xlscols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
    for i in range(26):
        for j in range(26):
            xlscols.append(xlscols[i] + xlscols[j])
    table = Table(
        displayName=table_name,
        ref=f'A1:{xlscols[len(table_headers)-1]}{len(table_rows) + 1}'
    )
    # Add a default style with striped rows and banded columns
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    ws.add_table(table)
   


if __name__ == "__main__":

    # get the current logged in username
    user = os.getlogin()
    print(f'Hello {user}!')

    # build and add a path to where the locally installed modules are commonly located
    new_path = f'//ihsnas1.net.inova.org/{user}/AppData/Roaming/Python/Python39/site-packages'
    print('Adding the following path for locally installed modules:')
    print('  ', new_path)
    if new_path not in sys.path:
        sys.path.append(new_path)

    # get the file headers
    header_lookup = get_csv_header_crosswalk('./settings/column_header_crosswalk.txt')
    print('Using the following column headers:')
    for k,v in header_lookup.items():
        print('  ', k, '=' , v)

    # get a list of all the input files
    input_directory = './input_files/'
    input_files = [d.strip() for d in os.listdir(input_directory) if d.endswith('.csv')]
    print('Using the following input files:')
    for f in input_files:
        print('  ', f)

    # load the data from the input files
    print('Loading input files...')
    data = load_input_files(input_directory, input_files, header_lookup)

    # calculate the reserve activity so that the reserve balances will roll-forward
    print('Calculating the reserve roll-foward activity for each account...')
    add_reserve_activity(data)

    # add a one-sentence description to each patient account
    print('Adding a one-sentence description to each patient account...')
    add_descriptions(data)

    #output_file = './output_files/output_file.xlsx'
    output_file = './output_files/output_file' + ' ' + datetime.today().isoformat(sep=' ', timespec='minutes').replace(':', '') + '.xlsx'

    # write the output to a new xlsx file
    print('Writing the data to the outputfile:')
    print('  ', output_file)
    write_to_file(output_file, data)

    print('################ DONE!!! ################')
    input('Press ENTER to continue...')

